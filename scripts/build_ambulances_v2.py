#!/usr/bin/env python3
"""Convert the partner's filtered-ambulance workbook into the app's own JSON.

Input : data/filtered_ambulances_v2.xlsx  (sheets "Applied Filters", "Ambulances")
Output: data/ambulances_v2.json

WHY A BUILD STEP AND NOT A LIVE XLSX READ
Reading .xlsx at request time would put openpyxl on the hot path of every map
draw for 142 rows that change roughly never. The JSON is committed, the build is
re-run by hand when the partner sends a new workbook.

WHAT THE DATA ACTUALLY IS — read this before using it
Every Unique ID appears EXACTLY ONCE, each stamped with one Day and one Time
Period. This is therefore a table of 142 OBSERVATIONS, not 142 ambulances each
carrying a weekly schedule. Filtering to a single day AND period leaves between
1 and 7 stations (48 of the 56 day x period cells are occupied at all), so a
reach analysis run on one slot legitimately shows most of Haryana uncovered.
That is the data being thin, not coverage collapsing, and the UI says so rather
than quietly drawing a red state.

Districts are normalised through districts.py so this file speaks the same
vocabulary as every other endpoint (SONIPAT, not Sonepat).
"""

from __future__ import annotations

import json
import os
import sys
from collections import Counter

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "dashboard"))

import districts  # noqa: E402  (needs the path insert above)

SRC = os.path.join(ROOT, "data", "filtered_ambulances_v2.xlsx")
OUT = os.path.join(ROOT, "data", "ambulances_v2.json")

# Chronological, not alphabetical — these drive a dropdown, and "03 AM" sorting
# before "12 AM" is the kind of detail that makes a UI feel broken.
PERIOD_ORDER = [
    "12 AM - 03 AM",
    "03 AM - 06 AM",
    "06 AM - 09 AM",
    "09 AM - 12 PM",
    "12 PM - 03 PM",
    "03 PM - 06 PM",
    "06 PM - 09 PM",
    "09 PM - 12 AM",
]
DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def main() -> int:
    if not os.path.exists(SRC):
        print(f"ERROR: missing {SRC}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(SRC, data_only=True)

    applied = {}
    if "Applied Filters" in wb.sheetnames:
        for k, v in wb["Applied Filters"].iter_rows(values_only=True):
            if k and str(k).strip().lower() != "filter":
                applied[str(k).strip()] = None if v is None else str(v).strip()

    rows = list(wb["Ambulances"].iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]

    out, skipped = [], 0
    for raw in rows[1:]:
        if raw[0] is None:
            continue
        r = dict(zip(header, raw))
        lat, lon = r.get("Latitude"), r.get("Longitude")
        # A station with no coordinate cannot be placed, reached from, or
        # measured against — dropping it is the only honest option, and the
        # count is reported so a silently shrinking file is impossible.
        if lat is None or lon is None:
            skipped += 1
            continue
        out.append(
            {
                "id": str(r["Unique ID"]).strip(),
                "day": (r.get("Day") or "").strip(),
                "period": (r.get("Time Period") or "").strip(),
                "district": districts.normalize(r.get("District")),
                "district_raw": (r.get("District") or "").strip(),
                "city": (r.get("City") or "").strip(),
                "postal_code": str(r.get("Postal Code") or "").strip(),
                "lat": float(lat),
                "lon": float(lon),
                "address": (r.get("Address") or "").strip(),
            }
        )
    return _write(out, applied, skipped)


def _write(out: list[dict], applied: dict, skipped: int) -> int:
    days = [d for d in DAY_ORDER if any(r["day"] == d for r in out)]
    days += sorted({r["day"] for r in out if r["day"] and r["day"] not in days})
    periods = [p for p in PERIOD_ORDER if any(r["period"] == p for r in out)]
    periods += sorted({r["period"] for r in out if r["period"] and r["period"] not in periods})

    # Per-slot counts ship with the file so the UI can warn BEFORE running an
    # analysis that only has two stations behind it.
    slot = Counter((r["day"], r["period"]) for r in out)

    payload = {
        "source": os.path.basename(SRC),
        "applied_filters": applied,
        "count": len(out),
        "skipped_no_coords": skipped,
        "days": days,
        "periods": periods,
        "districts": sorted({r["district"] for r in out if r["district"]}),
        "slot_counts": {f"{d}|{p}": n for (d, p), n in sorted(slot.items())},
        "ambulances": out,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=1, ensure_ascii=False)

    print(f"wrote {OUT}")
    print(f"  {len(out)} stations, {skipped} dropped for missing coordinates")
    print(f"  {len(days)} days x {len(periods)} periods, {len(slot)} slots occupied")
    if slot:
        print(f"  stations per slot: min {min(slot.values())}, max {max(slot.values())}")
    print(f"  {len(payload['districts'])} districts")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
